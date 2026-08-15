#!/usr/bin/env python3
"""Build a single Excel report from an evaluation.json produced by the resume-screener flow.

Usage:
    python build_report.py evaluation.json -o 简历筛选结果.xlsx

evaluation.json schema (see assets/evaluation_template.json):
{
  "jd_title": "...",           # optional
  "jd_breakdown": {            # optional but recommended
    "hard_requirements": ["..."],
    "bonus": ["..."],
    "ambiguous": ["..."]
  },
  "candidates": [
    {
      "file": "张三-产品.pdf",
      "name": "张三",
      "conclusion": "通过|待定|淘汰",
      "score": 82,
      "match_summary": "一句话总结",
      "evidence": [
        {
          "jd_requirement": "3年数据分析经验",
          "resume_snippet": "2021-2024 于XX任数据分析师",
          "assessment": "完全匹配|部分匹配|未提及"
        }
      ],
      "gaps": ["..."],
      "questions": ["..."],     # 待核实问题
      "interview_focus": ["..."],
      "notes": "..."
    }
  ]
}

The workbook contains three sheets:
  1. 筛选总表  - one row per candidate, sorted by score, color-coded conclusion
  2. 逐份点评  - detailed evidence / gaps / questions per candidate
  3. JD拆解    - the parsed JD requirements for transparency
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CONCLUSION_FILLS = {
    "通过": PatternFill("solid", fgColor="C6EFCE"),
    "待定": PatternFill("solid", fgColor="FFEB9C"),
    "淘汰": PatternFill("solid", fgColor="FFC7CE"),
}
CONCLUSION_FONTS = {
    "通过": Font(color="006100", bold=True),
    "待定": Font(color="9C6500", bold=True),
    "淘汰": Font(color="9C0006", bold=True),
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def fmt_assessment(assessment: str | None) -> str:
    if not assessment:
        return ""
    label = {"完全匹配": "完全匹配", "部分匹配": "部分匹配", "未提及": "未提及"}.get(
        assessment.strip(), assessment.strip()
    )
    return f"[{label}] "


def evidence_lines(evidence: list[dict]) -> str:
    lines = []
    for item in evidence or []:
        jd = (item.get("jd_requirement") or "").strip()
        snip = (item.get("resume_snippet") or "").strip()
        prefix = fmt_assessment(item.get("assessment"))
        line = prefix
        if jd:
            line += f"JD要求: {jd}"
        if jd and snip:
            line += "  ↔  "
        if snip:
            line += f"简历: {snip}"
        if line == prefix:
            line += "（未提供内容）"
        lines.append(line)
    return "\n".join(lines) if lines else "（无匹配证据）"


def join_lines(items) -> str:
    items = [str(x).strip() for x in (items or []) if str(x).strip()]
    return "\n".join(f"- {x}" for x in items) if items else ""


def estimate_height(texts: list[str], widths: list[float], min_height: float = 20.0) -> float:
    """Rough row-height estimate so wrapped CJK text is visible on open."""
    max_lines = 1
    for text, width in zip(texts, widths):
        if not text:
            continue
        chars_per_line = max(1, int(width / 2.1))
        lines = 0
        for seg in str(text).split("\n"):
            lines += max(1, math.ceil(len(seg) / chars_per_line))
        max_lines = max(max_lines, lines)
    return max(min_height, min(max_lines * 15 + 6, 409.0))


def style_header(ws, headers: list[str], widths: list[float]) -> None:
    for col, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"


def write_summary(ws, candidates: list[dict]) -> None:
    headers = ["序号", "结论", "匹配度", "候选人", "一句话点评", "主要亮点", "主要缺口", "待核实问题", "备注"]
    widths = [6, 8, 8, 18, 36, 48, 36, 36, 28]
    style_header(ws, headers, widths)

    for idx, cand in enumerate(candidates, start=1):
        name = cand.get("name") or cand.get("file") or "(未命名)"
        highlights = [
            f"{fmt_assessment(e.get('assessment'))}{e.get('jd_requirement') or e.get('resume_snippet')}"
            for e in (cand.get("evidence") or [])[:3]
            if e.get("jd_requirement") or e.get("resume_snippet")
        ]
        notes_parts = [cand.get("notes") or ""]
        if cand.get("interview_focus"):
            notes_parts.append("面试关注: " + "；".join(cand["interview_focus"]))
        row = [
            idx,
            cand.get("conclusion", ""),
            cand.get("score", ""),
            name,
            cand.get("match_summary", ""),
            "\n".join(highlights),
            join_lines(cand.get("gaps")),
            join_lines(cand.get("questions")),
            "\n".join(p for p in notes_parts if p),
        ]
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=idx + 1, column=col, value=value)
            cell.border = BORDER
            cell.alignment = WRAP
        ws.cell(row=idx + 1, column=1).alignment = CENTER
        ws.cell(row=idx + 1, column=2).alignment = CENTER
        ws.cell(row=idx + 1, column=3).alignment = CENTER
        conclusion = cand.get("conclusion", "")
        if conclusion in CONCLUSION_FILLS:
            ws.cell(row=idx + 1, column=2).fill = CONCLUSION_FILLS[conclusion]
            ws.cell(row=idx + 1, column=2).font = CONCLUSION_FONTS[conclusion]
        score = cand.get("score")
        if isinstance(score, (int, float)):
            ws.cell(row=idx + 1, column=3).number_format = "0" if score == int(score) else "0.0"
        ws.row_dimensions[idx + 1].height = estimate_height(
            [str(v) for v in row], widths
        )


def write_details(ws, candidates: list[dict]) -> None:
    headers = ["候选人", "结论", "匹配度", "JD要求 ↔ 简历证据", "缺口", "待核实问题", "面试关注点", "备注"]
    widths = [18, 8, 8, 70, 36, 36, 36, 28]
    style_header(ws, headers, widths)

    for idx, cand in enumerate(candidates, start=1):
        row = [
            cand.get("name") or cand.get("file") or "(未命名)",
            cand.get("conclusion", ""),
            cand.get("score", ""),
            evidence_lines(cand.get("evidence")),
            join_lines(cand.get("gaps")),
            join_lines(cand.get("questions")),
            join_lines(cand.get("interview_focus")),
            cand.get("notes") or "",
        ]
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=idx + 1, column=col, value=value)
            cell.border = BORDER
            cell.alignment = WRAP
        ws.cell(row=idx + 1, column=2).alignment = CENTER
        ws.cell(row=idx + 1, column=3).alignment = CENTER
        conclusion = cand.get("conclusion", "")
        if conclusion in CONCLUSION_FILLS:
            ws.cell(row=idx + 1, column=2).fill = CONCLUSION_FILLS[conclusion]
            ws.cell(row=idx + 1, column=2).font = CONCLUSION_FONTS[conclusion]
        score = cand.get("score")
        if isinstance(score, (int, float)):
            ws.cell(row=idx + 1, column=3).number_format = "0" if score == int(score) else "0.0"
        ws.row_dimensions[idx + 1].height = estimate_height([str(v) for v in row], widths)


def write_jd(ws, jd_breakdown: dict) -> None:
    headers = ["类型", "要求", "来源/说明"]
    widths = [14, 60, 60]
    style_header(ws, headers, widths)
    rows = []
    for key, label in (
        ("hard_requirements", "硬性要求"),
        ("bonus", "加分项"),
        ("ambiguous", "模糊项"),
    ):
        for item in jd_breakdown.get(key) or []:
            if isinstance(item, str):
                rows.append((label, item, ""))
            elif isinstance(item, dict):
                rows.append((label, item.get("requirement", ""), item.get("source", "")))
    if not rows:
        rows.append(("", "（未提供 JD 拆解）", ""))
    for idx, (kind, req, source) in enumerate(rows, start=1):
        for col, value in enumerate((kind, req, source), start=1):
            cell = ws.cell(row=idx + 1, column=col, value=value)
            cell.border = BORDER
            cell.alignment = WRAP
        ws.row_dimensions[idx + 1].height = estimate_height(
            [kind, req, source], widths
        )


def main() -> int:
    parser = argparse.ArgumentParser(description="Build resume-screening Excel report.")
    parser.add_argument("evaluation", help="Path to evaluation.json")
    parser.add_argument("-o", "--output", default="简历筛选结果.xlsx", help="Output xlsx path")
    args = parser.parse_args()

    if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    eval_path = Path(args.evaluation)
    if not eval_path.exists():
        print(f"[ERROR] 找不到文件: {eval_path}", file=sys.stderr)
        return 2
    data = json.loads(eval_path.read_text(encoding="utf-8"))
    candidates = data.get("candidates") or []
    if not isinstance(candidates, list):
        print("[ERROR] evaluation.json 缺少 candidates 数组", file=sys.stderr)
        return 2

    ordered = sorted(
        candidates, key=lambda c: c.get("score") if isinstance(c.get("score"), (int, float)) else -1,
        reverse=True,
    )

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "筛选总表"
    write_summary(ws_summary, ordered)
    write_details(wb.create_sheet("逐份点评"), ordered)
    write_jd(wb.create_sheet("JD拆解"), data.get("jd_breakdown") or {})

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))

    counts: dict[str, int] = {}
    for c in ordered:
        counts[c.get("conclusion", "")] = counts.get(c.get("conclusion", ""), 0) + 1
    print(f"已生成: {output}")
    print(f"候选人: {len(ordered)} 人；结论分布: {counts or '无'}")
    print("工作表: 筛选总表 / 逐份点评 / JD拆解")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
